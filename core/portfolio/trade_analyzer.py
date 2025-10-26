# trade_analyzer.py - Trade-Analyse und Excel-Export
import glob
import json
import os
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def parse_jsonl_logs(log_dir: str = "logs", run_timestamp: Optional[str] = None) -> List[Dict]:
    """
    Parst JSONL-Log-Dateien und extrahiert Trade-relevante Einträge.
    """
    trades = []

    # Finde relevante Log-Dateien
    if run_timestamp:
        # Spezifischer Run
        pattern = f"{log_dir}/run_{run_timestamp}_*.jsonl"
    else:
        # Neuester Run
        pattern = f"{log_dir}/run_*_trades.jsonl"
        files = glob.glob(pattern)
        if not files:
            pattern = f"{log_dir}/bot_log_*.jsonl"
            files = glob.glob(pattern)
        if files:
            # Sortiere nach Modifikationszeit und nimm neueste
            files.sort(key=os.path.getmtime, reverse=True)
            pattern = files[0].replace('\\', '/')

    # Parse alle gefundenen Dateien
    for log_file in glob.glob(pattern):
        try:
            with open(log_file, 'r', encoding='utf-8') as f:
                for line in f:
                    try:
                        entry = json.loads(line.strip())
                        # Filtere Trade-relevante Events
                        if entry.get('event_type') in [
                            'TRADE_OPENED', 'TRADE_CLOSED',
                            'TP_TO_SL_SWITCHED', 'BUY_ORDER_PLACED',
                            'BUY_ORDER_FILLED', 'SELL_LIMIT_IOC_EXECUTED',
                            'SELL_MARKET_EXECUTED', 'ORDER_FILLED_SYNC_EX'
                        ]:
                            trades.append(entry)
                    except json.JSONDecodeError:
                        continue
        except FileNotFoundError:
            continue

    return sorted(trades, key=lambda x: x.get('timestamp', ''))

def create_trade_summary(trades: List[Dict]) -> pd.DataFrame:
    """
    Erstellt eine strukturierte Trade-Zusammenfassung.
    """
    summary_data = []
    open_trades = {}

    for entry in trades:
        event_type = entry.get('event_type')
        symbol = entry.get('symbol')
        timestamp = entry.get('timestamp', '')

        if event_type == 'TRADE_OPENED':
            # Trade-Eröffnung
            open_trades[symbol] = {
                'symbol': symbol,
                'open_timestamp': timestamp,
                'entry_price': entry.get('entry_price', 0),
                'amount': entry.get('amount', 0),
                'cost_usdt': entry.get('cost_usdt', 0),
                'fee_usdt': entry.get('fee_usdt', 0),
                'tp_target': entry.get('tp_target', 0),
                'sl_target': entry.get('sl_target', 0),
                'tp_percentage': entry.get('tp_percentage', 0),
                'sl_percentage': entry.get('sl_percentage', 0),
                'status': 'OPEN',
                'switches': []
            }

        elif event_type == 'TP_TO_SL_SWITCHED' and symbol in open_trades:
            # TP→SL Switch
            open_trades[symbol]['switches'].append({
                'timestamp': timestamp,
                'current_price': entry.get('current_price', 0),
                'current_pnl_pct': entry.get('current_pnl_pct', 0)
            })
            open_trades[symbol]['switched_to_sl'] = True

        elif event_type == 'TRADE_CLOSED' and symbol in open_trades:
            # Trade-Schließung
            trade = open_trades[symbol]
            trade.update({
                'close_timestamp': timestamp,
                'exit_price': entry.get('exit_price', 0),
                'exit_type': entry.get('exit_type', ''),
                'exit_reason': entry.get('exit_reason', ''),
                'pnl_percentage': entry.get('pnl_percentage', 0),
                'profit_usdt': entry.get('profit_usdt', 0),
                'revenue_usdt': entry.get('revenue_usdt', 0),
                'status': 'CLOSED'
            })

            # Berechne Trade-Dauer
            try:
                open_dt = datetime.fromisoformat(trade['open_timestamp'].replace('Z', '+00:00'))
                close_dt = datetime.fromisoformat(trade['close_timestamp'].replace('Z', '+00:00'))
                duration = close_dt - open_dt
                trade['duration_minutes'] = duration.total_seconds() / 60
            except (ValueError, KeyError, TypeError):
                # Handle invalid timestamp formats or missing keys
                trade['duration_minutes'] = 0

            summary_data.append(trade)
            del open_trades[symbol]

    # Füge noch offene Trades hinzu
    for symbol, trade in open_trades.items():
        trade['status'] = 'OPEN'
        summary_data.append(trade)

    # Erstelle DataFrame
    df = pd.DataFrame(summary_data)

    # Sortiere nach Timestamp
    if not df.empty and 'open_timestamp' in df.columns:
        df = df.sort_values('open_timestamp')

    return df

def export_to_excel(df: pd.DataFrame, output_file: str = None):
    """
    Exportiert Trade-Summary als formatiertes Excel-File.
    """
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'trade_analysis_{timestamp}.xlsx'

    # Erstelle Workbook
    wb = Workbook()

    # === Sheet 1: Detaillierte Trades ===
    ws_trades = wb.active
    ws_trades.title = "Trade Details"

    # Header-Stil
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Spalten für Trade-Details
    trade_columns = [
        'symbol', 'status', 'open_timestamp', 'close_timestamp', 'duration_minutes',
        'entry_price', 'exit_price', 'amount', 'cost_usdt',
        'exit_type', 'pnl_percentage', 'profit_usdt', 'revenue_usdt',
        'tp_target', 'sl_target', 'switched_to_sl'
    ]

    # Filtere verfügbare Spalten
    available_columns = [col for col in trade_columns if col in df.columns]

    if not df.empty:
        # Schreibe Header
        for col_idx, col_name in enumerate(available_columns, 1):
            cell = ws_trades.cell(row=1, column=col_idx, value=col_name.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Schreibe Daten
        for row_idx, row in enumerate(df[available_columns].itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_trades.cell(row=row_idx, column=col_idx, value=value)

                # Formatierung basierend auf Spalte
                col_name = available_columns[col_idx-1]
                if 'price' in col_name or 'usdt' in col_name:
                    cell.number_format = '#,##0.00000000' if 'price' in col_name else '#,##0.00'
                elif 'percentage' in col_name or 'pnl' in col_name:
                    cell.number_format = '+0.00%;-0.00%'
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.font = Font(color="008000")  # Grün
                        elif value < 0:
                            cell.font = Font(color="FF0000")  # Rot
                elif col_name == 'status':
                    if value == 'CLOSED':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == 'OPEN':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif col_name == 'exit_type':
                    if value == 'TP':
                        cell.font = Font(color="008000", bold=True)
                    elif value == 'SL':
                        cell.font = Font(color="FF0000", bold=True)

        # Auto-Spaltenbreite
        for column_cells in ws_trades.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws_trades.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

    # === Sheet 2: Zusammenfassung ===
    ws_summary = wb.create_sheet("Summary")

    # Berechne Statistiken
    if not df.empty:
        closed_trades = df[df['status'] == 'CLOSED'] if 'status' in df.columns else df

        if not closed_trades.empty and 'profit_usdt' in closed_trades.columns:
            total_trades = len(closed_trades)
            winning_trades = len(closed_trades[closed_trades['profit_usdt'] > 0])
            losing_trades = len(closed_trades[closed_trades['profit_usdt'] < 0])

            total_profit = closed_trades['profit_usdt'].sum()
            avg_profit = closed_trades['profit_usdt'].mean()

            win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0

            avg_winner = closed_trades[closed_trades['profit_usdt'] > 0]['profit_usdt'].mean() if winning_trades > 0 else 0
            avg_loser = closed_trades[closed_trades['profit_usdt'] < 0]['profit_usdt'].mean() if losing_trades > 0 else 0

            # TP/SL Statistiken
            tp_exits = len(closed_trades[closed_trades['exit_type'] == 'TP']) if 'exit_type' in closed_trades.columns else 0
            sl_exits = len(closed_trades[closed_trades['exit_type'] == 'SL']) if 'exit_type' in closed_trades.columns else 0

            # Schreibe Summary
            summary_data = [
                ['Metric', 'Value'],
                ['Total Trades', total_trades],
                ['Winning Trades', winning_trades],
                ['Losing Trades', losing_trades],
                ['Win Rate', f'{win_rate:.2f}%'],
                ['', ''],
                ['Total Profit/Loss (USDT)', f'{total_profit:.2f}'],
                ['Average Profit/Loss (USDT)', f'{avg_profit:.2f}'],
                ['Average Winner (USDT)', f'{avg_winner:.2f}'],
                ['Average Loser (USDT)', f'{avg_loser:.2f}'],
                ['', ''],
                ['Take Profit Exits', tp_exits],
                ['Stop Loss Exits', sl_exits],
                ['Other Exits', total_trades - tp_exits - sl_exits]
            ]

            for row_idx, (metric, value) in enumerate(summary_data, 1):
                cell_metric = ws_summary.cell(row=row_idx, column=1, value=metric)
                cell_value = ws_summary.cell(row=row_idx, column=2, value=value)

                if row_idx == 1:
                    cell_metric.font = header_font
                    cell_metric.fill = header_fill
                    cell_value.font = header_font
                    cell_value.fill = header_fill
                elif metric in ['Total Profit/Loss (USDT)', 'Average Profit/Loss (USDT)']:
                    if isinstance(value, str) and '-' in value:
                        cell_value.font = Font(color="FF0000")
                    else:
                        cell_value.font = Font(color="008000")

            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 15

    # === Sheet 3: Performance Chart Data ===
    ws_perf = wb.create_sheet("Performance")

    if not df.empty and 'close_timestamp' in df.columns:
        closed_trades = df[df['status'] == 'CLOSED'].copy() if 'status' in df.columns else df.copy()

        if not closed_trades.empty and 'profit_usdt' in closed_trades.columns:
            # Kumulative P&L
            closed_trades = closed_trades.sort_values('close_timestamp')
            closed_trades['cumulative_pnl'] = closed_trades['profit_usdt'].cumsum()

            perf_data = closed_trades[['close_timestamp', 'symbol', 'profit_usdt', 'cumulative_pnl']]

            # Schreibe Header
            headers = ['Timestamp', 'Symbol', 'Trade P&L', 'Cumulative P&L']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_perf.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Schreibe Daten
            for row_idx, row in enumerate(perf_data.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws_perf.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx >= 3:  # P&L Spalten
                        cell.number_format = '#,##0.00'
                        if isinstance(value, (int, float)):
                            if value > 0:
                                cell.font = Font(color="008000")
                            elif value < 0:
                                cell.font = Font(color="FF0000")

    # Speichere Excel-Datei
    wb.save(output_file)
    return output_file

def analyze_current_run():
    """
    Analysiert den aktuellen Bot-Run und erstellt Excel-Report.
    """
    import logging
    logger = logging.getLogger(__name__)

    logger.info("Analysiere Trade-Logs...",
               extra={'event_type': 'TRADE_ANALYSIS_START'})

    # Parse Logs
    trades = parse_jsonl_logs()
    logger.info(f"Gefunden: {len(trades)} Trade-Events",
               extra={'event_type': 'TRADE_EVENTS_PARSED', 'count': len(trades)})

    # Erstelle Summary
    df = create_trade_summary(trades)
    logger.info(f"Verarbeitet: {len(df)} Trades",
               extra={'event_type': 'TRADE_SUMMARY_CREATED', 'count': len(df)})

    # Exportiere zu Excel
    output_file = export_to_excel(df)
    logger.info(f"Excel-Report erstellt: {output_file}",
               extra={'event_type': 'TRADE_REPORT_EXPORTED', 'file': output_file})

    return output_file

if __name__ == "__main__":
    analyze_current_run()
